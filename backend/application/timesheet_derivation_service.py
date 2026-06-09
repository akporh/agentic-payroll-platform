"""
Timesheet Derivation Application Service.

Orchestrates the three-step timesheet pipeline:
  1. upload_timesheet()   — parse Excel, validate, store timesheet_entry rows (PENDING)
  2. trigger_derivation() — run domain derivation per employee, store results (DERIVED / FAILED)
  3. approve_period()     — write payroll_input rows atomically, set entries APPROVED

This service has DB access. It delegates pure computation to:
  backend.domain.payroll.timesheet_derivation (no infra imports)

Hours-per-day defaults (configurable per workspace in a future sprint):
  DAY / S.DAY  → 8h
  2_SHIFT      → 8h
  4_SHIFT      → 12h
"""

from __future__ import annotations

import io
import json
import logging
from datetime import date, timedelta
from decimal import Decimal

from sqlalchemy import text

from backend.infra.db.session import SessionLocal
from backend.infra.repositories import (
    attendance_config_repo,
    timesheet_repo,
    payroll_input_repo,
)
from backend.infra.repositories.public_holiday_repo import get_effective_ph_list
from backend.domain.payroll.timesheet_derivation import (
    derive_payroll_inputs,
    is_numeric,
    DerivationSummary,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Hours-per-day defaults by shift type (overridable in workspace config later)
# ---------------------------------------------------------------------------
_DEFAULT_HOURS_PER_DAY: dict[str, Decimal] = {
    "DAY":     Decimal("8"),
    "S.DAY":   Decimal("8"),
    "2_SHIFT": Decimal("8"),
    "4_SHIFT": Decimal("12"),
}


def _get_hours_per_day(shift_type: str) -> Decimal:
    return _DEFAULT_HOURS_PER_DAY.get(shift_type, Decimal("8"))


def _count_working_days(start: date, end: date, ph_dates: set[date]) -> int:
    """Count Mon–Fri days in [start, end] excluding public holidays."""
    count = 0
    current = start
    while current <= end:
        if current.weekday() < 5 and current not in ph_dates:
            count += 1
        current += timedelta(days=1)
    return count


def _get_workspace_country(workspace_id: str) -> str | None:
    db = SessionLocal()
    try:
        row = db.execute(
            text("SELECT country_code FROM workspace WHERE workspace_id = :wid"),
            {"wid": workspace_id},
        ).fetchone()
        return row[0] if row else None
    finally:
        db.close()


def _get_employee_map(workspace_id: str) -> dict[str, dict]:
    """Return {employee_number: {employee_id, shift_type, full_name}} for active employees."""
    db = SessionLocal()
    try:
        rows = db.execute(
            text("""
                SELECT e.employee_id, e.employee_number, e.full_name,
                       ec.shift_type, ec.start_date, ec.end_date
                FROM employee e
                LEFT JOIN LATERAL (
                    SELECT ec2.*
                    FROM   employee_contract ec2
                    WHERE  ec2.employee_id = e.employee_id
                    ORDER  BY COALESCE(ec2.end_date, '9999-12-31') DESC,
                              ec2.start_date DESC NULLS LAST
                    LIMIT  1
                ) ec ON true
                WHERE e.workspace_id = :wid
                  AND e.status       = 'ACTIVE'
            """),
            {"wid": workspace_id},
        ).fetchall()

        return {
            str(r[1]): {
                "employee_id":    str(r[0]),
                "employee_number": str(r[1]),
                "full_name":      r[2] or str(r[1]),
                "shift_type":     r[3],
                "contract_start": r[4],
                "contract_end":   r[5],
            }
            for r in rows
            if r[1]
        }
    finally:
        db.close()


# ---------------------------------------------------------------------------
# Upload
# ---------------------------------------------------------------------------

def upload_timesheet(
    workspace_id: str,
    period_start: date,
    period_end: date,
    file_bytes: bytes,
) -> dict:
    """Parse an Excel timesheet file, validate, and store timesheet_entry rows.

    Returns an upload response dict:
    {
        employees_found: int,
        rows_accepted: int,
        rows_rejected: list[{employee_number, errors}],
        warnings: list[str],
    }
    """
    import openpyxl  # imported here so the rest of the module works without it

    MAX_UPLOAD_BYTES = 10 * 1024 * 1024  # 10 MB
    if len(file_bytes) > MAX_UPLOAD_BYTES:
        raise ValueError("Timesheet file exceeds 10 MB limit.")

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    employee_map = _get_employee_map(workspace_id)
    known_codes  = attendance_config_repo.get_all_code_client_codes(workspace_id)
    policies_map = attendance_config_repo.get_attendance_policies_for_derivation(workspace_id)

    # Parse column headers from row 1
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]

    try:
        emp_num_col  = headers.index("employee_number") + 1
        shift_type_col = headers.index("shift_type") + 1 if "shift_type" in headers else None
    except ValueError:
        return {
            "employees_found": 0,
            "rows_accepted":   0,
            "rows_rejected":   [],
            "warnings": ["Column 'employee_number' is required but not found in the uploaded file."],
        }

    # Build date column map: col_index → date (headers like "2026-01-21" or "2026-01-29*")
    ph_flagged_dates: set[date] = set()
    date_cols: dict[int, date] = {}
    for col_idx, hdr in enumerate(headers, start=1):
        if col_idx in (emp_num_col, shift_type_col):
            continue
        if hdr and isinstance(hdr, str) and len(hdr) >= 10:
            raw = str(hdr).strip()
            is_ph = raw.endswith("*")
            date_str = raw.rstrip("*").strip()
            try:
                d = date.fromisoformat(date_str)
                if period_start <= d <= period_end:
                    date_cols[col_idx] = d
                    if is_ph:
                        ph_flagged_dates.add(d)
            except ValueError:
                pass

    # Validate PH column dates against workspace PH config
    country_code = _get_workspace_country(workspace_id)
    ph_list = get_effective_ph_list(workspace_id, country_code, str(period_start), str(period_end))
    actual_ph_dates = {entry["date"] if isinstance(entry["date"], date) else date.fromisoformat(str(entry["date"])) for entry in ph_list}
    ph_header_mismatches = ph_flagged_dates - actual_ph_dates

    warnings: list[str] = []
    if ph_header_mismatches:
        for d in sorted(ph_header_mismatches):
            warnings.append(f"Column header '{d}*' is flagged as PH but this date is not in the workspace PH calendar.")

    # Prefetch APPROVED employee IDs for this period — re-upload must not overwrite approved evidence.
    approved_employee_ids = timesheet_repo.get_approved_employee_ids(workspace_id, period_start)

    rows_accepted = 0
    rows_rejected: list[dict] = []
    seen_employee_numbers: set[str] = set()

    for row_idx in range(2, ws.max_row + 1):
        emp_number = ws.cell(row_idx, emp_num_col).value
        if emp_number is None:
            continue
        emp_number = str(emp_number).strip()
        if not emp_number:
            continue

        row_errors: list[str] = []
        row_warnings: list[str] = []

        # Duplicate employee_number in same upload
        if emp_number in seen_employee_numbers:
            rows_rejected.append({"employee_number": emp_number, "errors": ["Duplicate employee_number in upload."]})
            continue
        seen_employee_numbers.add(emp_number)

        emp = employee_map.get(emp_number)
        if emp is not None and emp["employee_id"] in approved_employee_ids:
            rows_rejected.append({"employee_number": emp_number, "errors": ["Timesheet is APPROVED and cannot be overwritten."]})
            continue

        if emp is None:
            rows_rejected.append({"employee_number": emp_number, "errors": ["Employee not found in workspace."]})
            continue

        # shift_type validation
        shift_type = emp["shift_type"]
        if shift_type is None:
            rows_rejected.append({
                "employee_number": emp_number,
                "errors": ["shift_type is not configured — update the employee record and re-upload."],
            })
            continue

        # Build attendance grid from date columns
        grid: dict[str, str | int | float] = {}
        unknown_codes: list[str] = []
        inactive_code_warnings: list[str] = []
        no_policy_warnings: list[str] = []

        for col_idx, d in date_cols.items():
            cell_val = ws.cell(row_idx, col_idx).value
            if cell_val is None or cell_val == "":
                continue

            cell_str = str(cell_val).strip()
            date_key = d.isoformat()

            if is_numeric(cell_str):
                try:
                    val = float(cell_str)
                    if val < 0 or val > 24:
                        row_errors.append(f"{date_key}: numeric value {cell_str} out of range (0–24).")
                        continue
                    grid[date_key] = val
                except ValueError:
                    row_errors.append(f"{date_key}: could not parse numeric value {cell_str!r}.")
            else:
                code = cell_str
                if code not in known_codes:
                    unknown_codes.append(f"{date_key}: unknown code '{code}'")
                else:
                    policy = policies_map.get(code)
                    if policy is None:
                        no_policy_warnings.append(
                            f"Code '{code}' (cell {date_key}) has no policy configured — derivation will fail until resolved."
                        )
                    elif not policy.get("is_active", True):
                        inactive_code_warnings.append(
                            f"Code '{code}' (cell {date_key}) is inactive — contact your admin to re-enable it."
                        )
                    grid[date_key] = code

        if unknown_codes:
            row_errors.extend(unknown_codes)

        if row_errors:
            rows_rejected.append({"employee_number": emp_number, "errors": row_errors})
            continue

        # Warn but still store
        for w in no_policy_warnings + inactive_code_warnings:
            warnings.append(f"{emp_number}: {w}")

        timesheet_repo.upsert_timesheet_entry(
            workspace_id=workspace_id,
            employee_id=emp["employee_id"],
            period_start=period_start,
            period_end=period_end,
            grid_dict=grid,
        )
        rows_accepted += 1

    return {
        "employees_found": len(seen_employee_numbers),
        "rows_accepted":   rows_accepted,
        "rows_rejected":   rows_rejected,
        "warnings":        warnings,
    }


# ---------------------------------------------------------------------------
# Derivation
# ---------------------------------------------------------------------------

def trigger_derivation(
    workspace_id: str,
    period_start: date,
    period_end: date,
) -> dict:
    """Run derivation for all PENDING/FAILED entries for a workspace/period.

    Per-employee atomic: each employee's result is committed independently.
    Returns a summary of outcomes.
    """
    entries = timesheet_repo.get_entries_for_period(workspace_id, period_start)
    entries = [e for e in entries if e["derivation_status"] in ("PENDING", "FAILED")]

    if not entries:
        return {"message": "No PENDING or FAILED entries to process.", "processed": 0, "failed": 0}

    employee_map   = _get_employee_map(workspace_id)
    policies_map   = attendance_config_repo.get_attendance_policies_for_derivation(workspace_id)
    country_code   = _get_workspace_country(workspace_id)
    ph_list        = get_effective_ph_list(workspace_id, country_code, str(period_start), str(period_end))
    ph_dates: set[date] = {
        entry["date"] if isinstance(entry["date"], date) else date.fromisoformat(str(entry["date"]))
        for entry in ph_list
    }

    # OT trigger config from workspace_payroll_config
    db = SessionLocal()
    try:
        cfg_row = db.execute(
            text("""
                SELECT ph_rate_code
                FROM workspace_payroll_config
                WHERE workspace_id = :wid AND effective_from <= CURRENT_DATE
                ORDER BY effective_from DESC LIMIT 1
            """),
            {"wid": workspace_id},
        ).fetchone()
    finally:
        db.close()

    ot_trigger_config = [
        {"trigger_type": "EXCESS_HOURS",  "rate_code": "OT001"},
        {"trigger_type": "SATURDAY",      "rate_code": "OT002"},
        {"trigger_type": "PUBLIC_HOLIDAY", "rate_code": cfg_row[0] if cfg_row else "OT003"},
    ]

    processed = 0
    failed    = 0
    results   = []

    for entry in entries:
        emp_id     = entry["employee_id"]
        entry_id   = entry["timesheet_entry_id"]

        # Find employee by employee_id (reverse lookup)
        emp = next((v for v in employee_map.values() if v["employee_id"] == emp_id), None)
        if emp is None:
            timesheet_repo.update_derivation_result(
                workspace_id, entry_id, "FAILED", error="Employee not found or no active contract."
            )
            failed += 1
            continue

        shift_type = emp["shift_type"]
        if not shift_type:
            timesheet_repo.update_derivation_result(
                workspace_id, entry_id, "FAILED",
                error="shift_type is not configured — update the employee record."
            )
            failed += 1
            continue

        # Fetch full entry with grid
        full_entry = timesheet_repo.get_entry_with_grid(workspace_id, emp_id, period_start)
        if not full_entry:
            failed += 1
            continue

        grid = full_entry.get("attendance_grid_jsonb") or {}

        # Prefetch validation: all non-numeric codes must have a policy
        missing_policies = [
            code for code in set(str(v) for v in grid.values() if v and not is_numeric(v))
            if code not in policies_map
        ]
        if missing_policies:
            timesheet_repo.update_derivation_result(
                workspace_id, entry_id, "FAILED",
                error=f"No policy configured for codes: {', '.join(sorted(missing_policies))}"
            )
            failed += 1
            continue

        # Clamp contract window to period (H1)
        contract_start = emp.get("contract_start") or period_start
        contract_end   = emp.get("contract_end")   or period_end
        active_start   = max(period_start, contract_start) if isinstance(contract_start, date) else period_start
        active_end     = min(period_end,   contract_end)   if isinstance(contract_end, date)   else period_end

        hours_per_day   = _get_hours_per_day(shift_type)
        working_days    = _count_working_days(active_start, active_end, ph_dates)
        expected_hours  = Decimal(str(working_days)) * hours_per_day

        # Convert grid keys to dates for the domain function
        grid_by_date: dict[date, object] = {}
        for k, v in grid.items():
            try:
                grid_by_date[date.fromisoformat(str(k))] = v
            except ValueError:
                pass

        try:
            payroll_input_dicts, summary = derive_payroll_inputs(
                attendance_grid=grid_by_date,
                attendance_policies=policies_map,
                ot_trigger_config=ot_trigger_config,
                ph_date_set=ph_dates,
                shift_type=shift_type,
                contract_window=(active_start, active_end),
                hours_per_day=hours_per_day,
                expected_hours=expected_hours,
            )
        except Exception as exc:  # noqa: BLE001
            timesheet_repo.update_derivation_result(
                workspace_id, entry_id, "FAILED", error=str(exc)
            )
            failed += 1
            continue

        # Build policy snapshot for audit trail
        unique_codes = {str(v) for v in grid.values() if v and not is_numeric(v)}
        snapshot = {
            code: {
                "counts_as_paid":              policies_map[code]["counts_as_paid"],
                "counts_towards_ot_threshold": policies_map[code]["counts_towards_ot_threshold"],
                "hours_equivalent":            str(policies_map[code]["hours_equivalent"]) if policies_map[code]["hours_equivalent"] is not None else None,
                "unit_fraction":               str(policies_map[code]["unit_fraction"])    if policies_map[code]["unit_fraction"]    is not None else None,
            }
            for code in unique_codes
            if code in policies_map
        }

        # Encode Decimal values in payroll_input_dicts for JSON storage
        storable_inputs = [
            {**d, "quantity": float(d["quantity"])}
            for d in payroll_input_dicts
        ]

        derivation_summary_dict = {
            **summary.to_dict(),
            "payroll_input_dicts": storable_inputs,
            "ph_dates_used": [str(d) for d in sorted(ph_dates)],
        }

        timesheet_repo.update_derivation_result(
            workspace_id=workspace_id,
            timesheet_entry_id=entry_id,
            status="DERIVED",
            summary=derivation_summary_dict,
            snapshot=snapshot,
        )
        processed += 1
        results.append({"employee_id": emp_id, "status": "DERIVED", "proration_factor": float(summary.proration_factor)})

    return {"processed": processed, "failed": failed, "results": results}


# ---------------------------------------------------------------------------
# Approval
# ---------------------------------------------------------------------------

def approve_period(
    workspace_id: str,
    period_start: date,
    period_end: date,
) -> dict:
    """Approve all DERIVED entries for a period.

    Atomically writes payroll_input rows and sets derivation_status = APPROVED.
    Fails if any entry is not in DERIVED status.
    """
    entries = timesheet_repo.get_approved_entries_for_period(workspace_id, period_start)

    if not entries:
        # Check if there are non-DERIVED entries blocking approval
        all_entries = timesheet_repo.get_entries_for_period(workspace_id, period_start)
        non_derived = [e for e in all_entries if e["derivation_status"] != "DERIVED"]
        if non_derived:
            return {
                "approved": 0,
                "error": f"{len(non_derived)} employee(s) not in DERIVED status: "
                         + ", ".join(e["employee_name"] + f" ({e['derivation_status']})" for e in non_derived[:5]),
            }
        return {"approved": 0, "error": "No DERIVED entries found for this period."}

    db = SessionLocal()
    try:
        approved_count = 0
        for entry in entries:
            emp_id       = entry["employee_id"]
            summary      = entry.get("derivation_summary") or {}
            payroll_dicts = summary.get("payroll_input_dicts", [])

            # Delete any prior unclaimed TIMESHEET rows for this employee/period
            payroll_input_repo.delete_unclaimed_timesheet_inputs(
                workspace_id, emp_id, period_start, period_end
            )

            # Write new payroll_input rows (Decimal from float for insertion)
            rows_to_insert = [
                {
                    "employee_id":    emp_id,
                    "input_code":     d["input_code"],
                    "input_category": d["input_category"],
                    "quantity":       Decimal(str(d["quantity"])),
                    "reference_date": None,
                }
                for d in payroll_dicts
                if d.get("quantity") and float(d["quantity"]) > 0
            ]

            if rows_to_insert:
                payroll_input_repo.batch_create_timesheet_inputs(workspace_id, rows_to_insert, db=db)

            approved_count += 1

        # Set all DERIVED entries to APPROVED in one sweep
        timesheet_repo.set_entries_approved(workspace_id, period_start, db=db)
        db.commit()

    except Exception:
        db.rollback()
        raise
    finally:
        db.close()

    return {"approved": approved_count}
